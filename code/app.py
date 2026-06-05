#!/usr/bin/env python3
"""
app.py — FastAPI backend for the Amoeba Morphometrics Pipeline web interface.

Run with:
    cd code && uvicorn app:app --reload --port 8000
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Any, AsyncGenerator

import numpy as np
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    FileResponse,
    JSONResponse,
    Response,
    StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from PIL import Image

# ── Paths ─────────────────────────────────────────────────────────────────────

CODE_DIR    = Path(__file__).parent          # /pipeline/code/
REPO_DIR    = CODE_DIR.parent                # /pipeline/
STATIC_DIR  = CODE_DIR / "static"
SCRIPTS_DIR = CODE_DIR / "scripts"
MODELS_DIR  = CODE_DIR / "models"
DATA_DIR     = REPO_DIR / "data"
INPUT_DIR    = DATA_DIR / "input"
CURATED_DIR  = DATA_DIR / "curated"
TRAINING_DIR = DATA_DIR / "training"
RESULTS_DIR = REPO_DIR / "results"

CONFIG_FILE         = REPO_DIR / "config.json"
CURATION_STATE_FILE = REPO_DIR / "curation_state.json"

IMAGE_EXTENSIONS = {".tif", ".tiff", ".jpg", ".jpeg", ".png"}

# In-memory caches — keyed by (analysis_id, seg_path) to avoid cross-analysis collisions
_mask_cache: dict[str, np.ndarray] = {}
_jobs: dict[str, dict[str, Any]] = {}
_env_status_cache: dict = {}  # cached result of environment check

# ── Environment helpers ───────────────────────────────────────────────────────

def find_conda() -> str | None:
    """Return path to conda/mamba/micromamba, checking PATH then common install dirs."""
    for cmd in ["mamba", "micromamba", "conda"]:
        found = shutil.which(cmd)
        if found:
            return found
    home = Path.home()
    for base in [
        home / "opt/miniconda3",  home / "opt/mambaforge",  home / "opt/miniforge3",
        home / "miniconda3",      home / "mambaforge",       home / "miniforge3",
        home / "anaconda3",       Path("/opt/conda"),        Path("/opt/miniconda3"),
        Path("/usr/local/miniconda3"),
    ]:
        for binary in ["bin/mamba", "bin/conda"]:
            p = base / binary
            if p.exists():
                return str(p)
    return None


def get_pipeline_python() -> str:
    """Return the best available Python for running pipeline scripts."""
    cfg = get_config()
    if cfg.get("python_path"):
        p = Path(cfg["python_path"])
        if p.exists():
            return str(p)
    home = Path.home()
    for env_name in ["morpheus", "morpheus-env", "napari-env"]:
        for conda_base in [
            home / "opt/miniconda3", home / "opt/mambaforge", home / "opt/miniforge3",
            home / "miniconda3",     home / "mambaforge",      home / "miniforge3",
            home / "anaconda3",      Path("/opt/conda"),        Path("/opt/miniconda3"),
        ]:
            p = conda_base / "envs" / env_name / "bin" / "python"
            if p.exists():
                return str(p)
    return sys.executable


# ── FastAPI app ───────────────────────────────────────────────────────────────

app = FastAPI(title="Amoeba Morphometrics Pipeline")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── State helpers ─────────────────────────────────────────────────────────────

def _read_json(path: Path, default: Any) -> Any:
    if path.exists():
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return default


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def get_config() -> dict:
    defaults: dict = {
        "strains": [],
        "analyses": [],
        "objectives": [],
    }
    cfg = _read_json(CONFIG_FILE, {})
    return {**defaults, **cfg}


def save_config(cfg: dict) -> None:
    _write_json(CONFIG_FILE, cfg)


def get_curation_state() -> dict:
    return _read_json(CURATION_STATE_FILE, {})


def measurements_path(analysis_id: str) -> Path:
    return RESULTS_DIR / analysis_id / "measurements.json"


# ── Environment check & install endpoints ────────────────────────────────────

REQUIRED_PACKAGES = ["cellpose", "napari", "tifffile", "pandas", "numpy", "scikit-image"]

@app.get("/api/env/status")
def api_env_status():
    global _env_status_cache
    python = get_pipeline_python()
    check_script = (
        "import json, sys\n"
        "pkgs = {}\n"
        "for name, imp in [\n"
        "  ('cellpose','cellpose'),('napari','napari'),('tifffile','tifffile'),\n"
        "  ('pandas','pandas'),('numpy','numpy'),('scikit-image','skimage'),\n"
        "  ('matplotlib','matplotlib'),('fastapi','fastapi'),('uvicorn','uvicorn'),\n"
        "]:\n"
        "  try:\n"
        "    m=__import__(imp); pkgs[name]=str(getattr(m,'__version__',getattr(m,'version','?')))\n"
        "  except: pkgs[name]=None\n"
        "print(json.dumps({'pyver':sys.version.split()[0],'packages':pkgs}))\n"
    )
    try:
        r = subprocess.run(
            [python, "-c", check_script],
            capture_output=True, text=True, timeout=30,
            env={**os.environ, "PYTHONUNBUFFERED": "1"},
        )
        if r.returncode == 0:
            data = json.loads(r.stdout.strip())
            ok = all(data["packages"].get(p) for p in REQUIRED_PACKAGES)
            result = {"ok": ok, "python": python, **data, "conda": find_conda()}
            _env_status_cache = result
            return result
    except Exception as e:
        pass
    return {"ok": False, "python": python, "packages": {}, "pyver": None, "conda": find_conda()}


@app.post("/api/env/set-python")
async def api_env_set_python(body: dict):
    path = body.get("python_path", "").strip()
    if path and not Path(path).exists():
        raise HTTPException(status_code=400, detail="Python path does not exist")
    cfg = get_config()
    cfg["python_path"] = path or None
    save_config(cfg)
    _env_status_cache.clear()
    return {"ok": True}


@app.post("/api/env/install/start")
async def api_env_install_start(body: dict):
    method = body.get("method", "conda")   # "conda" | "pip"
    env_name = body.get("env_name", "morpheus-env")
    conda = find_conda()
    env_yml = REPO_DIR / "environment.yml"

    if method == "conda":
        if not conda:
            raise HTTPException(status_code=400, detail="conda/mamba not found — use pip install or install Miniconda first")
        if env_yml.exists():
            # Preferred: create from environment.yml
            cmd = [conda, "env", "create", "-f", str(env_yml), "--force"]
        else:
            # Fallback: install packages directly
            cmd = [conda, "create", "-n", env_name, "-c", "conda-forge",
                   "python=3.11", "napari", "pyqt", "tifffile", "pandas",
                   "numpy", "scikit-image", "matplotlib", "openpyxl", "-y"]
    else:
        # pip fallback — installs into the current environment
        cmd = [sys.executable, "-m", "pip", "install",
               "cellpose>=4.0", "napari", "tifffile", "pandas", "numpy",
               "scikit-image", "matplotlib", "openpyxl", "fastapi", "uvicorn[standard]"]

    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {
        "cmd": cmd,
        "status": "pending",
        "lines": [],
        "type": "install",
        "env_name": env_name,
        "method": method,
    }
    return {"job_id": job_id}


@app.get("/api/env/install/events/{job_id}")
async def api_env_install_events(job_id: str):
    if job_id not in _jobs:
        raise HTTPException(status_code=404)

    async def stream():
        job = _jobs[job_id]
        job["status"] = "running"
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"

        def _send(data: dict) -> bytes:
            return f"data: {json.dumps(data)}\n\n".encode()

        try:
            proc = subprocess.Popen(
                job["cmd"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, env=env,
            )
            while True:
                line = await asyncio.get_event_loop().run_in_executor(None, proc.stdout.readline)
                if not line:
                    break
                line = line.rstrip()
                job["lines"].append(line)
                yield _send({"status": "running", "message": line})
                await asyncio.sleep(0)
            proc.wait()

            if proc.returncode == 0:
                # Try to find the new python and save it
                if job.get("method") == "conda":
                    new_py = get_pipeline_python()
                    if new_py != sys.executable:
                        cfg = get_config()
                        cfg["python_path"] = new_py
                        save_config(cfg)
                job["status"] = "done"
                yield _send({"status": "done",
                             "message": "Installation complete. Reload the page to re-check environment."})
            else:
                job["status"] = "error"
                yield _send({"status": "error", "message": f"Process exited with code {proc.returncode}"})
        except Exception as e:
            job["status"] = "error"
            yield _send({"status": "error", "message": str(e)})

    return StreamingResponse(stream(), media_type="text/event-stream")


@app.get("/api/env/ome-info")
def api_env_ome_info(strain: str = Query(...)):
    """Check first TIFF in strain input dir for embedded OME pixel size."""
    input_dir = INPUT_DIR / strain
    if not input_dir.exists():
        return {"is_ome": False, "pixel_um": None}
    for f in sorted(input_dir.rglob("*")):
        if f.suffix.lower() not in {".tif", ".tiff"}:
            continue
        try:
            import tifffile
            import xml.etree.ElementTree as ET
            with tifffile.TiffFile(str(f)) as tif:
                if tif.is_ome and tif.ome_metadata:
                    root = ET.fromstring(tif.ome_metadata)
                    for ns in [
                        "http://www.openmicroscopy.org/Schemas/OME/2016-06",
                        "http://www.openmicroscopy.org/Schemas/OME/2015-01",
                        "http://www.openmicroscopy.org/Schemas/OME/2013-10-dev-4",
                        "",
                    ]:
                        prefix = f"{{{ns}}}" if ns else ""
                        px = root.find(f".//{prefix}Pixels")
                        if px is not None:
                            x = px.get("PhysicalSizeX")
                            if x:
                                unit = px.get("PhysicalSizeXUnit", "µm")
                                val = float(x)
                                # Convert nm → µm if needed
                                if unit in ("nm", "nanometer"):
                                    val /= 1000
                                return {"is_ome": True, "pixel_um": round(val, 6), "filename": f.name}
                    return {"is_ome": True, "pixel_um": None, "filename": f.name}
        except Exception:
            pass
        break  # only check first TIFF
    return {"is_ome": False, "pixel_um": None}


# ── Config / Strain endpoints ─────────────────────────────────────────────────

@app.get("/api/config")
def api_get_config():
    return get_config()


@app.post("/api/config/strain/add")
async def api_add_strain(body: dict):
    name = (body.get("name") or "").strip()
    source_dir = (body.get("source_dir") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    if source_dir and not Path(source_dir).exists():
        raise HTTPException(status_code=400, detail=f"source_dir not found: {source_dir}")

    cfg = get_config()
    strains: list = cfg.get("strains", [])
    if any(s["name"] == name for s in strains):
        raise HTTPException(status_code=400, detail=f"Strain '{name}' already exists")

    strains.append({"name": name, "source_dir": source_dir, "color": "#4ade80"})
    cfg["strains"] = strains
    save_config(cfg)
    return {"strains": strains}


@app.delete("/api/config/strain/{name}")
def api_delete_strain(name: str):
    cfg = get_config()
    strains = [s for s in cfg.get("strains", []) if s["name"] != name]
    cfg["strains"] = strains
    save_config(cfg)

    removed_dirs = []

    # data/input/<strain>/
    input_dir = INPUT_DIR / name
    if input_dir.exists():
        shutil.rmtree(str(input_dir))
        removed_dirs.append(str(input_dir))

    # data/training/<strain>/
    training_dir = TRAINING_DIR / name
    if training_dir.exists():
        shutil.rmtree(str(training_dir))
        removed_dirs.append(str(training_dir))

    # data/curated/<analysis>/<strain>/ — remove from every analysis
    if CURATED_DIR.exists():
        for analysis_dir in CURATED_DIR.iterdir():
            if not analysis_dir.is_dir():
                continue
            strain_dir = analysis_dir / name
            if strain_dir.exists():
                shutil.rmtree(str(strain_dir))
                removed_dirs.append(str(strain_dir))

    return {"strains": strains, "removed_dirs": removed_dirs}


@app.get("/api/config/strain/import/{name}/stream")
async def api_import_strain_stream(name: str):
    """SSE stream: emits progress events as each image is copied."""
    cfg = get_config()
    strain = next((s for s in cfg.get("strains", []) if s["name"] == name), None)
    if strain is None:
        raise HTTPException(status_code=404, detail=f"Strain '{name}' not in config")

    source_dir = _resolve_dir(strain.get("source_dir", ""))
    if not source_dir or not Path(source_dir).exists():
        raise HTTPException(status_code=400, detail=f"source_dir not set or not found: {source_dir}")

    src = Path(source_dir)
    dest = INPUT_DIR / name
    dest.mkdir(parents=True, exist_ok=True)

    # Collect all images first so we know the total
    images = [f for f in src.rglob("*") if f.suffix.lower() in IMAGE_EXTENSIONS]
    total = len(images)

    async def generate():
        yield f"data: {json.dumps({'total': total, 'copied': 0, 'status': 'starting'})}\n\n"
        copied, errors = 0, []
        for img in images:
            rel = img.relative_to(src)
            target = dest / rel
            target.parent.mkdir(parents=True, exist_ok=True)
            try:
                await asyncio.get_event_loop().run_in_executor(
                    None, lambda s=img, t=target: shutil.copy2(str(s), str(t))
                )
                copied += 1
            except Exception as e:
                errors.append(str(e))
            yield f"data: {json.dumps({'total': total, 'copied': copied, 'filename': img.name, 'status': 'copying'})}\n\n"
        yield f"data: {json.dumps({'total': total, 'copied': copied, 'errors': errors, 'status': 'done'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.post("/api/config/strain/import-all")
async def api_import_all_strains():
    cfg = get_config()
    results = []
    for strain in cfg.get("strains", []):
        name = strain["name"]
        source_dir = _resolve_dir(strain.get("source_dir", ""))
        if not source_dir or not Path(source_dir).exists():
            results.append({"name": name, "copied": 0, "error": "source_dir not found"})
            continue
        src = Path(source_dir)
        dest = INPUT_DIR / name
        dest.mkdir(parents=True, exist_ok=True)
        copied, errors = 0, []
        for img in src.rglob("*"):
            if img.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            rel = img.relative_to(src)
            target = dest / rel
            target.parent.mkdir(parents=True, exist_ok=True)
            try:
                shutil.copy2(str(img), str(target))
                copied += 1
            except Exception as e:
                errors.append(str(e))
        results.append({"name": name, "copied": copied, "errors": errors})
    return {"results": results}


@app.get("/api/config/strain/status")
def api_strain_status():
    cfg = get_config()
    result = []
    for strain in cfg.get("strains", []):
        name = strain["name"]
        source_dir = _resolve_dir(strain.get("source_dir", ""))

        source_count = 0
        if source_dir and Path(source_dir).exists():
            source_count = sum(
                1 for f in Path(source_dir).rglob("*")
                if f.suffix.lower() in IMAGE_EXTENSIONS
            )

        imported_count = 0
        imported_dir = INPUT_DIR / name
        if imported_dir.exists():
            imported_count = sum(
                1 for f in imported_dir.rglob("*")
                if f.suffix.lower() in IMAGE_EXTENSIONS
            )

        result.append({
            "name": name,
            "source_dir": source_dir,
            "color": strain.get("color", "#4ade80"),
            "source_count": source_count,
            "imported_count": imported_count,
        })
    return result


# ── Analysis management ───────────────────────────────────────────────────────

@app.get("/api/analyses")
def api_get_analyses():
    cfg = get_config()
    return cfg.get("analyses", [])


@app.post("/api/analyses/add")
async def api_add_analysis(body: dict):
    analysis_id = (body.get("id") or "").strip()
    name = (body.get("name") or "").strip()
    if not analysis_id:
        raise HTTPException(status_code=400, detail="id required")
    if not name:
        raise HTTPException(status_code=400, detail="name required")

    cfg = get_config()
    analyses: list = cfg.get("analyses", [])
    if any(a["id"] == analysis_id for a in analyses):
        raise HTTPException(status_code=400, detail=f"Analysis id '{analysis_id}' already exists")

    analysis = {
        "id":             analysis_id,
        "name":           name,
        "model_path":     body.get("model_path", ""),
        "measurements":   body.get("measurements", [
            "length_um", "breadth_um", "aspect_ratio", "area_um2",
            "feret_max_um", "feret_min_um", "feret_aspect_ratio",
            "solidity", "perimeter_um",
        ]),
        "min_area":       int(body.get("min_area", 300)),
        "max_area":       int(body.get("max_area", 500000)),
        "diameter":       body.get("diameter"),
        "pixel_size_um":  body.get("pixel_size_um", 0.1075),
        "pixels_per_um":  body.get("pixels_per_um"),
        "strain_models":  body.get("strain_models", {}),
    }
    analyses.append(analysis)
    cfg["analyses"] = analyses
    save_config(cfg)
    return analysis


@app.patch("/api/analyses/{analysis_id}")
async def api_update_analysis(analysis_id: str, body: dict):
    """Update any fields of an existing analysis (except id)."""
    cfg = get_config()
    analyses = cfg.get("analyses", [])
    analysis = next((a for a in analyses if a["id"] == analysis_id), None)
    if analysis is None:
        raise HTTPException(status_code=404, detail=f"Analysis '{analysis_id}' not found")

    updatable = ["name", "model_path", "measurements", "min_area", "max_area",
                 "diameter", "pixel_size_um", "pixels_per_um"]
    for k in updatable:
        if k in body:
            analysis[k] = body[k]

    cfg["analyses"] = analyses
    save_config(cfg)
    return analysis


@app.patch("/api/analyses/{analysis_id}/strain-models")
async def api_set_strain_models(analysis_id: str, body: dict):
    """Set per-strain model overrides for an analysis. body: {strain_models: {name: path}}"""
    cfg = get_config()
    analyses = cfg.get("analyses", [])
    analysis = next((a for a in analyses if a["id"] == analysis_id), None)
    if analysis is None:
        raise HTTPException(status_code=404, detail=f"Analysis '{analysis_id}' not found")
    analysis["strain_models"] = body.get("strain_models", {})
    cfg["analyses"] = analyses
    save_config(cfg)
    return analysis


@app.delete("/api/analyses/{analysis_id}")
def api_delete_analysis(analysis_id: str):
    cfg = get_config()
    analyses = [a for a in cfg.get("analyses", []) if a["id"] != analysis_id]
    cfg["analyses"] = analyses
    save_config(cfg)
    return {"analyses": analyses}


# ── Objectives endpoints ──────────────────────────────────────────────────────

@app.get("/api/objectives")
def api_get_objectives():
    return get_config().get("objectives", [])


@app.post("/api/objectives/save")
async def api_save_objectives(body: dict):
    objectives = body.get("objectives", [])
    cfg = get_config()
    cfg["objectives"] = objectives
    save_config(cfg)
    return {"objectives": objectives}


@app.get("/api/measurements/available")
def api_measurements_available():
    # Ensure pipeline package is on sys.path
    code_dir = str(CODE_DIR)
    if code_dir not in sys.path:
        sys.path.insert(0, code_dir)
    try:
        from pipeline.measurements import metadata
        return metadata()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load measurements: {e}")


# ── Native directory picker ───────────────────────────────────────────────────

@app.post("/api/pick-directory")
def api_pick_directory():
    """Open native macOS file dialog via subprocess (tkinter needs its own main thread)."""
    script = (
        "import tkinter as tk, tkinter.filedialog as fd, json, sys\n"
        "root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)\n"
        "path = fd.askdirectory(title='Select Directory')\n"
        "root.destroy()\n"
        "print(json.dumps({'path': path or None}))\n"
    )
    try:
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode == 0 and result.stdout.strip():
            return json.loads(result.stdout.strip())
        raise HTTPException(status_code=500, detail=result.stderr or "Dialog failed")
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=408, detail="Dialog timed out (120 s)")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File dialog error: {e}")


# ── Models ────────────────────────────────────────────────────────────────────

@app.get("/api/models")
def api_models():
    models = []
    if MODELS_DIR.exists():
        for entry in sorted(MODELS_DIR.iterdir()):
            if entry.name.startswith(".") or entry.name == "README.md":
                continue
            if entry.is_file() or entry.is_dir():
                models.append({"name": entry.name, "path": str(entry)})
    return {"models": models}


# ── Image serving ─────────────────────────────────────────────────────────────

@app.get("/api/image")
def api_image(path: str = Query(...)):
    code_dir = str(CODE_DIR)
    if code_dir not in sys.path:
        sys.path.insert(0, code_dir)
    from pipeline.image_io import read_image, to_display_jpeg
    p = Path(path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    try:
        img, _ = read_image(p)
        data = to_display_jpeg(img)
        return Response(content=data, media_type="image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/thumbnail")
def api_thumbnail(path: str = Query(...)):
    code_dir = str(CODE_DIR)
    if code_dir not in sys.path:
        sys.path.insert(0, code_dir)
    from pipeline.image_io import make_thumbnail, read_image
    p = Path(path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    try:
        img, _ = read_image(p)
        data = make_thumbnail(img)
        return Response(content=data, media_type="image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Napari launcher ───────────────────────────────────────────────────────────

_napari_procs: dict[str, subprocess.Popen] = {}


@app.post("/api/launch/selector")
async def api_launch_selector(body: dict):
    """Launch select_images.py in napari browsing data/input/."""
    analysis_id = body.get("analysis_id", "default")
    destination = body.get("destination", "curated")   # 'training' or 'curated'

    if not INPUT_DIR.exists() or not any(INPUT_DIR.iterdir()):
        raise HTTPException(status_code=400, detail="data/input/ is empty — import strain images first")

    script = SCRIPTS_DIR / "select_images.py"
    if not script.exists():
        raise HTTPException(status_code=500, detail="select_images.py not found")

    sel_file = CODE_DIR / f"selections_{analysis_id}_{destination}.json"

    env = os.environ.copy()
    env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    env["QT_QPA_PLATFORM_PLUGIN_PATH"] = (
        str(Path(sys.executable).parent.parent
            / "lib/python3.11/site-packages/PyQt6/Qt6/plugins/platforms")
    )
    env["SELECTIONS_FILE"] = str(sel_file)

    proc = subprocess.Popen(
        [sys.executable, str(script), str(INPUT_DIR),
         "--target", str(body.get("target", 30))],
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    pid = str(proc.pid)
    _napari_procs[pid] = proc
    return {
        "pid": pid,
        "analysis_id": analysis_id,
        "destination": destination,
        "selections_file": str(sel_file),
    }


@app.get("/api/launch/status/{pid}")
def api_launch_status(pid: str):
    proc = _napari_procs.get(pid)
    if proc is None:
        return {"running": False, "pid": pid}
    running = proc.poll() is None
    return {"running": running, "pid": pid, "returncode": proc.returncode}


@app.post("/api/copy-to-set")
async def api_copy_to_set(body: dict):
    """Copy selections from a selections JSON into data/curated/<analysis_id>/."""
    analysis_id = body.get("analysis_id", "default")
    destination = body.get("destination", "curated")
    sel_file_path = body.get("selections_file") or str(
        CODE_DIR / f"selections_{analysis_id}_{destination}.json"
    )

    sel_data = _read_json(Path(sel_file_path), {"selected": []})
    selected_paths = sel_data.get("selected", [])

    if not selected_paths:
        return {
            "copied": 0,
            "errors": [],
            "message": "No selections found — open napari and select images first",
        }

    if destination == "training":
        dest_root = TRAINING_DIR          # data/training/<strain>/
    else:
        dest_root = CURATED_DIR / analysis_id  # data/curated/<analysis>/<strain>/

    copied, errors = 0, []

    for src_str in selected_paths:
        src = Path(src_str)
        if not src.exists():
            errors.append({"path": src_str, "error": "file not found"})
            continue
        # Preserve strain sub-directory (relative to INPUT_DIR)
        try:
            rel = src.relative_to(INPUT_DIR)
            strain = rel.parts[0]
        except ValueError:
            strain = src.parent.name
        dest_dir = dest_root / strain
        dest_dir.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(str(src), str(dest_dir / src.name))
            copied += 1
        except Exception as e:
            errors.append({"path": src_str, "error": str(e)})

    return {"copied": copied, "errors": errors, "destination": str(dest_root)}


@app.post("/api/use-all-images")
async def api_use_all_images(body: dict):
    """Copy all images from data/input/ into data/curated/<analysis_id>/."""
    analysis_id = body.get("analysis_id", "default")
    if not INPUT_DIR.exists():
        raise HTTPException(status_code=400, detail="data/input/ not found")

    dest_root = CURATED_DIR / analysis_id
    copied = 0
    for img in INPUT_DIR.rglob("*"):
        if img.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        rel = img.relative_to(INPUT_DIR)
        dest = dest_root / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        if not dest.exists():
            shutil.copy2(str(img), str(dest))
            copied += 1

    return {"copied": copied, "destination": str(dest_root)}


# ── Browser-based image selection (replaces napari select_images.py) ──────────

@app.get("/api/select/images")
def api_select_images(analysis_id: str = Query(...), destination: str = Query("curated")):
    """List all images in data/input/ annotated with selection/viewed state."""
    sel_file = CODE_DIR / f"selections_{analysis_id}_{destination}.json"
    sel_data = _read_json(sel_file, {"selected": [], "viewed": []})
    selected_abs = set(sel_data.get("selected", []))
    viewed_abs   = set(sel_data.get("viewed",   []))

    images = []
    if INPUT_DIR.exists():
        for f in sorted(INPUT_DIR.rglob("*")):
            if f.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            try:
                rel_to_input = f.relative_to(INPUT_DIR)
                strain = rel_to_input.parts[0] if len(rel_to_input.parts) > 1 else f.parent.name
            except ValueError:
                strain = f.parent.name
            abs_str = str(f)
            rel_str = str(f.relative_to(REPO_DIR))
            images.append({
                "path":      rel_str,
                "abs_path":  abs_str,
                "strain":    strain,
                "filename":  f.name,
                "selected":  abs_str in selected_abs,
                "viewed":    abs_str in viewed_abs,
            })

    selected_count = sum(1 for i in images if i["selected"])
    viewed_count   = sum(1 for i in images if i["viewed"])
    return {
        "images":          images,
        "selected_count":  selected_count,
        "viewed_count":    viewed_count,
        "total":           len(images),
        "selections_file": str(sel_file),
    }


@app.post("/api/select/save")
async def api_select_save(body: dict):
    """Persist browser image selections to the selections JSON file."""
    analysis_id = body.get("analysis_id", "default")
    destination = body.get("destination", "curated")
    selected    = body.get("selected", [])
    viewed      = body.get("viewed",   [])
    sel_file = CODE_DIR / f"selections_{analysis_id}_{destination}.json"
    data = {"selected": selected, "viewed": viewed, "count": len(selected)}
    sel_file.write_text(json.dumps(data, indent=2))
    return {"ok": True, "path": str(sel_file), "count": len(selected)}


@app.get("/api/set-status")
def api_set_status(analysis_id: str = Query(default="default")):
    """Return image counts in data/input/ and data/curated/<analysis_id>/ per strain."""
    cfg = get_config()
    strain_names = [s["name"] for s in cfg.get("strains", [])]

    # Fallback: discover strains from input dir if config is empty
    if not strain_names and INPUT_DIR.exists():
        strain_names = [
            d.name for d in sorted(INPUT_DIR.iterdir())
            if d.is_dir() and not d.name.startswith(".")
        ]

    result: dict = {"input": {}, "curated": {}}

    for strain in strain_names:
        n_input = 0
        input_dir = INPUT_DIR / strain
        if input_dir.exists():
            n_input = sum(
                1 for f in input_dir.rglob("*") if f.suffix.lower() in IMAGE_EXTENSIONS
            )
        if n_input > 0:
            result["input"][strain] = n_input

    # Curated counts: discover from disk so deletion of a config strain
    # never wipes the display for remaining strains still on disk.
    curated_root = CURATED_DIR / analysis_id
    if curated_root.exists():
        for strain_dir in sorted(curated_root.iterdir()):
            if not strain_dir.is_dir() or strain_dir.name.startswith("."):
                continue
            n_curated = sum(
                1 for f in strain_dir.rglob("*") if f.suffix.lower() in IMAGE_EXTENSIONS
            )
            if n_curated > 0:
                result["curated"][strain_dir.name] = n_curated

    return result


# ── Model training ────────────────────────────────────────────────────────────

@app.get("/api/training/strains")
def api_training_strains():
    """List strains that have images in data/training/, with their first image path."""
    strains = []
    if not TRAINING_DIR.exists():
        return strains
    for strain_dir in sorted(TRAINING_DIR.iterdir()):
        if not strain_dir.is_dir() or strain_dir.name.startswith("."):
            continue
        images = sorted(
            f for f in strain_dir.rglob("*")
            if f.suffix.lower() in IMAGE_EXTENSIONS
        )
        if images:
            strains.append({
                "name": strain_dir.name,
                "count": len(images),
                "first_image": str(images[0]),
            })
    return strains


@app.post("/api/launch/cellpose-gui")
def api_launch_cellpose_gui(body: dict = {}):
    """Launch Cellpose GUI, optionally opening the first image of a training strain."""
    image_path: str | None = body.get("image_path")

    env = os.environ.copy()
    env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    env["QT_QPA_PLATFORM_PLUGIN_PATH"] = str(
        Path(sys.executable).parent.parent
        / "lib/python3.11/site-packages/PyQt6/Qt6/plugins/platforms"
    )

    # Launch via gui.run(image=...) in a subprocess so the image loads on startup
    image_arg = f", image={repr(str(image_path))}" if image_path else ""
    script = (
        f"from cellpose.gui import gui; gui.run({image_arg.lstrip(', ')})"
        if image_path
        else "from cellpose.gui import gui; gui.run()"
    )
    proc = subprocess.Popen([sys.executable, "-c", script], env=env)
    return {"pid": proc.pid, "message": "Cellpose GUI launched", "image": image_path}


@app.post("/api/training/prepare")
async def api_training_prepare(body: dict):
    """Collect masked image pairs from data/training/ into a flat masks dir.
    body: { strains: ["StrainA", "StrainB"] }  — empty list = all strains
    """
    selected_strains: list[str] = body.get("strains", [])
    script = SCRIPTS_DIR / "prepare_training.py"
    if not script.exists():
        raise HTTPException(status_code=500, detail="prepare_training.py not found")

    masks_dir = DATA_DIR / "training_masks"
    cmd = [
        sys.executable, str(script),
        "--src-dir", str(TRAINING_DIR),
        "--out-dir", str(masks_dir),
    ]
    if selected_strains:
        cmd += ["--strains", ",".join(selected_strains)]

    job_id = str(uuid.uuid4())[:8]
    env = os.environ.copy()
    env["KMP_DUPLICATE_LIB_OK"] = "TRUE"

    _jobs[job_id] = {
        "cmd": cmd,
        "status": "pending",
        "lines": [],
        "type": "prepare",
    }
    return {"job_id": job_id}


@app.post("/api/training/start")
async def api_training_start(body: dict):
    analysis_id = body.get("analysis_id", "default")
    train_masks_dir = DATA_DIR / "training_masks"

    if not train_masks_dir.exists():
        raise HTTPException(
            status_code=400,
            detail=f"Training masks not found at {train_masks_dir} — run Prepare first",
        )

    base_model = body.get("base_model", "cyto2")
    epochs = int(body.get("epochs", 200))
    model_name = body.get("model_name", "CustomModel")

    cmd = [
        sys.executable, "-m", "cellpose",
        "--train",
        "--dir", str(train_masks_dir),
        "--pretrained_model", base_model,
        "--mask_filter", "_seg.npy",
        "--min_train_masks", "1",
        "--n_epochs", str(epochs),
        "--model_name_out", model_name,
        "--verbose",
    ]

    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {
        "cmd": cmd,
        "status": "pending",
        "lines": [],
        "type": "training",
        "model_name": model_name,
        "train_masks_dir": str(train_masks_dir),
    }
    return {"job_id": job_id}


# ── Measurement / SSE job ─────────────────────────────────────────────────────

@app.post("/api/measure/start")
async def api_measure_start(body: dict):
    analysis_id = body.get("analysis_id", "default")

    cfg = get_config()
    analysis = next((a for a in cfg.get("analyses", []) if a["id"] == analysis_id), None)
    if analysis is None:
        raise HTTPException(status_code=404, detail=f"Analysis '{analysis_id}' not found in config")

    default_model = _resolve_model_path(analysis.get("model_path", ""))
    strain_models = {k: _resolve_model_path(v) for k, v in analysis.get("strain_models", {}).items()}

    curated_dir = CURATED_DIR / analysis_id
    if not curated_dir.exists():
        raise HTTPException(
            status_code=400,
            detail=f"Curated set not found at {curated_dir} — run Select Images first",
        )

    # Collect strain dirs present in the curated set
    strain_dirs = [d for d in sorted(curated_dir.iterdir()) if d.is_dir()]
    if not strain_dirs:
        raise HTTPException(status_code=400, detail="Curated set is empty — import and select images first")

    # Build one command per strain (allows per-strain model override)
    out_dir = RESULTS_DIR / analysis_id
    out_dir.mkdir(parents=True, exist_ok=True)

    script = SCRIPTS_DIR / "measure_training_set.py"
    if not script.exists():
        raise HTTPException(status_code=500, detail="measure_training_set.py not found")

    def _build_cmd(strain_name: str, model: str) -> list[str]:
        cmd = [
            sys.executable, str(script),
            "--model",    str(model),
            "--dir",      str(curated_dir),   # analysis dir; script iterates strain subdirs
            "--strain",   strain_name,
            "--outdir",   str(out_dir),
            "--min_area", str(analysis.get("min_area", 300)),
            "--max_area", str(analysis.get("max_area", 500000)),
        ]
        if analysis.get("diameter") is not None:
            cmd += ["--diameter", str(analysis["diameter"])]
        if analysis.get("pixels_per_um") is not None:
            cmd += ["--pixels_per_um", str(analysis["pixels_per_um"])]
        elif analysis.get("pixel_size_um") is not None:
            cmd += ["--pixel_size", str(analysis["pixel_size_um"])]
        return cmd

    cmds: list[tuple[str, list[str]]] = []  # [(strain_name, cmd)]
    missing_model_strains = []
    for sd in strain_dirs:
        model = strain_models.get(sd.name) or default_model
        if not model:
            missing_model_strains.append(sd.name)
            continue
        cmds.append((sd.name, _build_cmd(sd.name, model)))

    if missing_model_strains:
        raise HTTPException(
            status_code=400,
            detail=f"No model assigned for strains: {', '.join(missing_model_strains)}. "
                   "Set a default model on the analysis or a per-strain override.",
        )

    if not cmds:
        raise HTTPException(status_code=400, detail="No strains to measure")

    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {
        "cmds": cmds,   # list of (strain, cmd) — executed sequentially
        "status": "pending",
        "lines": [],
        "analysis_id": analysis_id,
    }
    return {"job_id": job_id}


@app.get("/api/measure/events/{job_id}")
async def api_measure_events(job_id: str):
    if job_id not in _jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    async def event_stream() -> AsyncGenerator[bytes, None]:
        job = _jobs[job_id]
        job["status"] = "running"

        # Support both old single-cmd jobs and new multi-strain cmds
        cmds: list[tuple[str, list[str]]] = job.get("cmds") or [("", job.get("cmd", []))]
        n_strains = len(cmds)

        env = os.environ.copy()
        env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
        env["PYTHONUNBUFFERED"] = "1"  # flush print() output immediately

        # Lines to suppress from the log (noisy Cellpose deprecation warnings)
        _SUPPRESS = ("channels deprecated", "diam_mean", "DeprecationWarning",
                     "WARNING", "UserWarning")

        def _send(data: dict) -> bytes:
            return f"data: {json.dumps(data)}\n\n".encode()

        overall_progress = 0
        overall_total = 0

        try:
            for strain_idx, (strain_name, cmd) in enumerate(cmds):
                prefix = f"[{strain_name}] " if strain_name else ""
                yield _send({
                    "status": "running",
                    "message": f"{prefix}Starting ({strain_idx + 1}/{n_strains})…",
                    "progress": overall_progress,
                    "total": max(overall_total, 1),
                })

                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    cwd=str(CODE_DIR),
                    env=env,
                )
                strain_total = 0
                strain_progress = 0

                while True:
                    line = await asyncio.get_event_loop().run_in_executor(
                        None, proc.stdout.readline
                    )
                    if not line:
                        break
                    line = line.rstrip()
                    job["lines"].append(line)

                    # Skip noisy deprecation/warning lines from Cellpose
                    if any(s in line for s in _SUPPRESS):
                        await asyncio.sleep(0)
                        continue

                    # Parse [X/Y] progress anywhere in the line
                    m = re.search(r"\[\s*(\d+)\s*/\s*(\d+)\]", line)
                    if m:
                        strain_progress = int(m.group(1))
                        strain_total    = int(m.group(2))
                        overall_total   = max(overall_total, strain_total)
                        overall_progress = strain_progress

                    yield _send({
                        "status": "running",
                        "message": f"{prefix}{line}",
                        "progress": overall_progress,
                        "total": max(overall_total, 1),
                    })
                    await asyncio.sleep(0)

                proc.wait()
                rc = proc.returncode
                if rc != 0:
                    job["status"] = "error"
                    yield _send({
                        "status": "error",
                        "message": f"{prefix}Process exited with code {rc}",
                        "progress": overall_progress,
                        "total": max(overall_total, 1),
                    })
                    return

            job["status"] = "done"

            # After training: move output model into models/ dir
            if job.get("type") == "training":
                mname = job.get("model_name", "CustomModel")
                masks_dir = Path(job.get("train_masks_dir", ""))
                # Cellpose saves to <train_dir>/models/<model_name_out>
                trained = masks_dir / "models" / mname
                if trained.exists():
                    MODELS_DIR.mkdir(parents=True, exist_ok=True)
                    dest = MODELS_DIR / mname
                    shutil.move(str(trained), str(dest))
                    done_msg = f"Training complete. Model saved to models/{mname}"
                else:
                    done_msg = f"Training complete. (Model not found at {trained} — check Cellpose output)"
            else:
                done_msg = f"Measurement complete — {n_strains} strain(s) processed."

            yield _send({
                "status": "done",
                "message": done_msg,
                "progress": overall_total,
                "total": overall_total,
            })

        except Exception as e:
            job["status"] = "error"
            yield _send({"status": "error", "message": str(e), "progress": 0, "total": 0})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── Curation helpers ──────────────────────────────────────────────────────────

def _resolve_dir(p: str) -> str:
    """Make a directory path absolute. Relative paths are resolved against REPO_DIR."""
    if not p:
        return p
    path = Path(p)
    return p if path.is_absolute() else str(REPO_DIR / path)

def _resolve_model_path(p: str) -> str:
    """Resolve a model path stored in config: absolute paths used as-is; bare names
    (e.g. 'test_model') are resolved to MODELS_DIR/<name>."""
    if not p:
        return p
    path = Path(p)
    if path.is_absolute():
        return p
    # Bare name or relative — resolve from MODELS_DIR
    return str(MODELS_DIR / path.name)

_PATH_FIELDS      = ("filepath", "seg_path")
_CELL_PATH_FIELDS = ("crop_path", "feret_crop_path")

def _resolve_path(p: str) -> str:
    """Make a path absolute. If already absolute and exists, return as-is.
    Otherwise treat as relative to REPO_DIR (for portable test-dataset JSON)."""
    if not p:
        return p
    path = Path(p)
    if path.is_absolute():
        return p
    return str(REPO_DIR / path)

def _load_measurements(analysis_id: str) -> list[dict]:
    mp = measurements_path(analysis_id)
    if not mp.exists():
        raise HTTPException(
            status_code=404,
            detail=f"measurements.json not found for analysis '{analysis_id}' — run Measure step first",
        )
    with open(mp) as f:
        data = json.load(f)
    # Resolve relative paths so all downstream code works with absolute paths
    for img in data:
        for field in _PATH_FIELDS:
            if field in img:
                img[field] = _resolve_path(img[field])
        for cell in img.get("cells", []):
            for field in _CELL_PATH_FIELDS:
                if field in cell:
                    cell[field] = _resolve_path(cell[field])
    return data


def _mask_cache_key(analysis_id: str, seg_path: str) -> str:
    return f"{analysis_id}::{seg_path}"


def _load_mask_cached(analysis_id: str, seg_path: str) -> np.ndarray:
    key = _mask_cache_key(analysis_id, seg_path)
    if key not in _mask_cache:
        from skimage import segmentation
        data = np.load(seg_path, allow_pickle=True).item()
        mask = data["masks"].astype(np.int32)
        _mask_cache[key] = segmentation.clear_border(mask)
    return _mask_cache[key]


def _curation_key(analysis_id: str, filename: str, cell_id: int | str) -> str:
    return f"{analysis_id}:{filename}:{cell_id}"


# ── Curation endpoints ────────────────────────────────────────────────────────

@app.get("/api/curation/images")
def api_curation_images(analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    return [
        {
            "idx":      idx,
            "strain":   img["strain"],
            "filename": img["filename"],
            "filepath": img["filepath"],
            "n_cells":  len(img.get("cells", [])),
        }
        for idx, img in enumerate(images_data)
    ]


@app.get("/api/curation/image/{idx}")
def api_curation_image(idx: int, analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    if idx < 0 or idx >= len(images_data):
        raise HTTPException(status_code=404, detail="Index out of range")

    img_data = images_data[idx]
    curation = get_curation_state()
    filename = img_data["filename"]

    cells = []
    for cell in img_data.get("cells", []):
        key = _curation_key(analysis_id, filename, cell["cell_id"])
        selected = curation.get(key, True)
        cells.append({**cell, "selected": selected})

    return {
        "idx":           idx,
        "strain":        img_data["strain"],
        "filename":      filename,
        "filepath":      img_data["filepath"],
        "seg_path":      img_data.get("seg_path", ""),
        "pixel_size_um": img_data.get("pixel_size_um", 0.1075),
        "cells":         cells,
    }


@app.get("/api/curation/overlay/{idx}")
def api_curation_overlay(idx: int, analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    if idx < 0 or idx >= len(images_data):
        raise HTTPException(status_code=404, detail="Index out of range")

    img_data = images_data[idx]
    seg_path = img_data.get("seg_path", "")
    if not seg_path or not Path(seg_path).exists():
        raise HTTPException(status_code=404, detail="Segmentation mask not found")

    curation = get_curation_state()
    filename = img_data["filename"]
    measured_ids = {c["cell_id"] for c in img_data.get("cells", [])}

    mask = _load_mask_cached(analysis_id, seg_path)
    h, w = mask.shape
    overlay = np.zeros((h, w, 4), dtype=np.uint8)

    for cell in img_data.get("cells", []):
        cid = cell["cell_id"]
        if cid not in measured_ids:
            continue
        key = _curation_key(analysis_id, filename, cid)
        is_selected = curation.get(key, True)
        region = mask == cid
        if is_selected:
            overlay[region] = [0, 220, 80, 160]
        else:
            overlay[region] = [220, 60, 60, 160]

    pil_img = Image.fromarray(overlay, mode="RGBA")
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png")


@app.post("/api/curation/cell-at")
async def api_curation_cell_at(body: dict):
    idx = body.get("idx")
    x = body.get("x")
    y = body.get("y")
    analysis_id = body.get("analysis_id", "default")
    if idx is None or x is None or y is None:
        raise HTTPException(status_code=400, detail="idx, x, y required")

    images_data = _load_measurements(analysis_id)
    if idx < 0 or idx >= len(images_data):
        raise HTTPException(status_code=404, detail="Index out of range")

    img_data = images_data[idx]
    seg_path = img_data.get("seg_path", "")
    if not seg_path or not Path(seg_path).exists():
        raise HTTPException(status_code=404, detail="Mask not found")

    mask = _load_mask_cached(analysis_id, seg_path)
    row = int(round(y))
    col = int(round(x))
    h, w = mask.shape
    if row < 0 or row >= h or col < 0 or col >= w:
        return {"cell_id": 0}

    return {"cell_id": int(mask[row, col])}


@app.post("/api/curation/toggle")
async def api_curation_toggle(body: dict):
    filename = body.get("filename", "")
    cell_id = body.get("cell_id")
    analysis_id = body.get("analysis_id", "default")
    if not filename or cell_id is None:
        raise HTTPException(status_code=400, detail="filename and cell_id required")

    key = _curation_key(analysis_id, filename, cell_id)
    curation = get_curation_state()
    current = curation.get(key, True)
    curation[key] = not current
    _write_json(CURATION_STATE_FILE, curation)
    return {"selected": curation[key]}


@app.post("/api/curation/accept-all/{idx}")
async def api_curation_accept_all(idx: int, analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    if idx < 0 or idx >= len(images_data):
        raise HTTPException(status_code=404, detail="Index out of range")
    img_data = images_data[idx]
    filename = img_data["filename"]
    curation = get_curation_state()
    for cell in img_data.get("cells", []):
        key = _curation_key(analysis_id, filename, cell["cell_id"])
        curation[key] = True
    _write_json(CURATION_STATE_FILE, curation)
    seg_path = img_data.get("seg_path", "")
    _mask_cache.pop(_mask_cache_key(analysis_id, seg_path), None)
    return {"ok": True}


@app.post("/api/curation/reject-all/{idx}")
async def api_curation_reject_all(idx: int, analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    if idx < 0 or idx >= len(images_data):
        raise HTTPException(status_code=404, detail="Index out of range")
    img_data = images_data[idx]
    filename = img_data["filename"]
    curation = get_curation_state()
    for cell in img_data.get("cells", []):
        key = _curation_key(analysis_id, filename, cell["cell_id"])
        curation[key] = False
    _write_json(CURATION_STATE_FILE, curation)
    seg_path = img_data.get("seg_path", "")
    _mask_cache.pop(_mask_cache_key(analysis_id, seg_path), None)
    return {"ok": True}


@app.get("/api/curation/export")
def api_curation_export(analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    curation = get_curation_state()

    # Gather column names dynamically from first cell that has data
    columns = ["strain", "filename", "cell_id", "pixel_size_um", "filepath"]
    for img in images_data:
        for cell in img.get("cells", []):
            for k in cell.keys():
                if k not in columns:
                    columns.append(k)
            break
        if len(columns) > 5:
            break

    # Ensure standard columns are present
    standard = [
        "length_um", "breadth_um", "aspect_ratio",
        "feret_max_um", "feret_min_um", "feret_aspect_ratio",
        "area_um2", "solidity", "perimeter_um",
    ]
    for col in standard:
        if col not in columns:
            columns.append(col)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()

    for img_data in images_data:
        filename = img_data["filename"]
        pixel_size = img_data.get("pixel_size_um", 0.1075)
        for cell in img_data.get("cells", []):
            key = _curation_key(analysis_id, filename, cell["cell_id"])
            if not curation.get(key, True):
                continue
            row = {
                "strain":        img_data["strain"],
                "filename":      filename,
                "pixel_size_um": pixel_size,
                "filepath":      img_data["filepath"],
            }
            row.update(cell)
            writer.writerow(row)

    # Save to results dir
    out_dir = RESULTS_DIR / analysis_id
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "curated_cells.csv"
    csv_path.write_text(buf.getvalue())

    return Response(
        content=buf.getvalue().encode(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=curated_cells_{analysis_id}.csv"},
    )


# ── Curate v2: cell-grid endpoints ───────────────────────────────────────────

@app.get("/api/curation/strains")
def api_curation_strains(analysis_id: str = Query(default="default")):
    images_data = _load_measurements(analysis_id)
    strains = sorted({img["strain"] for img in images_data})
    return strains


@app.get("/api/curation/cells")
def api_curation_cells(analysis_id: str = Query(default="default"), strain: str = Query(default="")):
    images_data = _load_measurements(analysis_id)
    curation = get_curation_state()
    cells = []
    for img in images_data:
        if strain and img["strain"] != strain:
            continue
        for cell in img.get("cells", []):
            key = _curation_key(analysis_id, img["filename"], cell["cell_id"])
            raw = curation.get(key, True)
            if raw is True:    morph = "accepted"
            elif raw is False: morph = "rejected"
            else:              morph = str(raw)
            cells.append({
                "filename": img["filename"],
                "strain": img["strain"],
                "cell_id": cell["cell_id"],
                "crop_path": cell.get("crop_path", ""),
                "pixel_size_um": cell.get("pixel_size_um", img.get("pixel_size_um", 0.1075)),
                "morphotype": morph,
                "length_um": cell.get("length_um"),
                "breadth_um": cell.get("breadth_um"),
                "aspect_ratio": cell.get("aspect_ratio"),
                "feret_max_um": cell.get("feret_max_um"),
                "feret_min_um": cell.get("feret_min_um"),
                "area_um2": cell.get("area_um2"),
                "solidity": cell.get("solidity"),
                "centroid_y_px": cell.get("centroid_y_px"),
                "centroid_x_px": cell.get("centroid_x_px"),
                "orientation_rad":      cell.get("orientation_rad"),
                "feret_max_angle_rad":  cell.get("feret_max_angle_rad"),
                "feret_min_angle_rad":  cell.get("feret_min_angle_rad"),
                "bbox":                 cell.get("bbox"),
            })
    return cells


@app.get("/api/curation/file")
def api_curation_file(path: str = Query(...), thumb: bool = Query(False)):
    p = Path(path)
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    # Allow serving from RESULTS_DIR (crops) or DATA_DIR (source images)
    try:
        p.relative_to(RESULTS_DIR)
    except ValueError:
        try:
            p.relative_to(DATA_DIR)
        except ValueError:
            raise HTTPException(status_code=403, detail="Forbidden")
    suffix = p.suffix.lower()
    if suffix in (".tif", ".tiff"):
        # Browsers can't render TIFF — convert to normalised 8-bit PNG in memory
        import io, tifffile, numpy as np
        img = tifffile.imread(str(p))
        if img.ndim == 3:
            img = img[0] if img.shape[0] <= 4 else img[..., 0]
        lo, hi = np.percentile(img, [1, 99])
        img8 = ((np.clip(img.astype(np.float32), lo, hi) - lo) /
                (hi - lo + 1e-6) * 255).astype(np.uint8)
        from PIL import Image as PILImage
        import io as _io
        pil_img = PILImage.fromarray(img8)
        if thumb:
            pil_img.thumbnail((280, 280), PILImage.LANCZOS)
        buf = _io.BytesIO()
        pil_img.save(buf, format="PNG")
        return Response(content=buf.getvalue(), media_type="image/png")
    if thumb:
        from PIL import Image as PILImage
        import io as _io
        pil_img = PILImage.open(str(p)).convert("L")
        pil_img.thumbnail((280, 280), PILImage.LANCZOS)
        buf = _io.BytesIO()
        pil_img.save(buf, format="PNG")
        return Response(content=buf.getvalue(), media_type="image/png")
    media_type = "image/png" if suffix == ".png" else "image/jpeg"
    return Response(content=p.read_bytes(), media_type=media_type)


_DEFAULT_MORPHOTYPES = [
    {"id": "accepted", "name": "Accepted", "color": "#4a6830"},
    {"id": "rejected", "name": "Rejected", "color": "#8b2020"},
]


@app.get("/api/curation/morphotypes")
def api_morphotypes_get(analysis_id: str = Query(default="default")):
    cfg = get_config()
    analysis = next((a for a in cfg.get("analyses", []) if a["id"] == analysis_id), None)
    if analysis is None:
        return _DEFAULT_MORPHOTYPES
    return analysis.get("morphotypes", _DEFAULT_MORPHOTYPES)


@app.post("/api/curation/morphotypes")
async def api_morphotypes_save(body: dict):
    analysis_id = body.get("analysis_id", "default")
    morphotypes = body.get("morphotypes", _DEFAULT_MORPHOTYPES)
    cfg = get_config()
    for a in cfg.get("analyses", []):
        if a["id"] == analysis_id:
            a["morphotypes"] = morphotypes
            break
    save_config(cfg)
    return {"ok": True}


@app.post("/api/curation/assign")
async def api_curation_assign(body: dict):
    analysis_id = body.get("analysis_id", "default")
    assignments = body.get("assignments", [])  # [{filename, cell_id, morphotype}]
    curation = get_curation_state()
    for a in assignments:
        key = _curation_key(analysis_id, a["filename"], a["cell_id"])
        curation[key] = a["morphotype"]
    _write_json(CURATION_STATE_FILE, curation)
    return {"ok": True, "count": len(assignments)}


@app.post("/api/curation/export-split")
async def api_curation_export_split(body: dict):
    analysis_id = body.get("analysis_id", "default")
    images_data = _load_measurements(analysis_id)
    curation = get_curation_state()
    out_base = RESULTS_DIR / analysis_id / "morphotypes"
    counts: dict[str, int] = {}
    for img in images_data:
        for cell in img.get("cells", []):
            key = _curation_key(analysis_id, img["filename"], cell["cell_id"])
            raw = curation.get(key, True)
            morph = "accepted" if raw is True else ("rejected" if raw is False else str(raw))
            crop_path = cell.get("crop_path", "")
            if not crop_path or not Path(crop_path).exists():
                continue
            dest_dir = out_base / morph
            dest_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(crop_path, dest_dir / Path(crop_path).name)
            counts[morph] = counts.get(morph, 0) + 1
    return {"exported": counts, "out_dir": str(out_base)}


# ── Test dataset ──────────────────────────────────────────────────────────────

_TEST_STRAIN = {
    "name": "Nolandella",
    "source_dir": "data/input/Nolandella",
    "color": "#4ade80",
}
_TEST_ANALYSIS = {
    "id": "nolandella_test",
    "name": "Nolandella Test",
    "model_path": "test_model",
    "measurements": [
        "length_um", "breadth_um", "aspect_ratio", "area_um2",
        "perimeter_um", "solidity", "feret_max_um", "feret_min_um", "feret_aspect_ratio",
    ],
    "min_area": 300,
    "max_area": 500000,
    "diameter": None,
    "pixel_size_um": 0.1075,
    "pixels_per_um": None,
    "strain_models": {},
}


@app.get("/api/test-data/status")
def api_test_data_status():
    cfg = get_config()
    has_strain   = any(s["name"] == "Nolandella" for s in cfg.get("strains", []))
    has_analysis = any(a["id"] == "nolandella_test" for a in cfg.get("analyses", []))
    mpath = measurements_path("nolandella_test")
    has_measurements = mpath.exists()
    has_files = (CURATED_DIR / "nolandella_test").exists() and (INPUT_DIR / "Nolandella").exists()
    n_cells = 0
    if has_measurements:
        try:
            with open(mpath) as f:
                data = json.load(f)
            n_cells = sum(len(img.get("cells", [])) for img in data)
        except Exception:
            pass
    return {
        "active": has_strain and has_analysis and has_measurements,
        "has_files": has_files,
        "n_cells": n_cells,
    }


@app.post("/api/test-data/load")
async def api_test_data_load():
    mpath = measurements_path("nolandella_test")
    if not mpath.exists():
        raise HTTPException(
            status_code=400,
            detail="Test measurements not found — the repo may be missing test data files."
        )
    cfg = get_config()
    strains = cfg.get("strains", [])
    if not any(s["name"] == "Nolandella" for s in strains):
        strains.insert(0, dict(_TEST_STRAIN))
    cfg["strains"] = strains
    analyses = cfg.get("analyses", [])
    if not any(a["id"] == "nolandella_test" for a in analyses):
        analyses.insert(0, dict(_TEST_ANALYSIS))
    cfg["analyses"] = analyses
    save_config(cfg)
    return {"ok": True}


@app.delete("/api/test-data")
def api_test_data_delete():
    cfg = get_config()
    cfg["strains"]  = [s for s in cfg.get("strains",  []) if s["name"] != "Nolandella"]
    cfg["analyses"] = [a for a in cfg.get("analyses", []) if a["id"]   != "nolandella_test"]
    save_config(cfg)

    curation = get_curation_state()
    curation = {k: v for k, v in curation.items() if not k.startswith("nolandella_test:")}
    _write_json(CURATION_STATE_FILE, curation)

    removed = []
    for path in [
        RESULTS_DIR / "nolandella_test",
        CURATED_DIR / "nolandella_test",
        INPUT_DIR   / "Nolandella",
    ]:
        if path.exists():
            shutil.rmtree(str(path))
            removed.append(str(path))

    return {"ok": True, "removed": removed}


# ── Results ───────────────────────────────────────────────────────────────────

def _build_curated_df(analysis_id: str):
    """Return a DataFrame of non-rejected cells with a morphotype column."""
    import pandas as pd

    curation = get_curation_state()

    def _morph(raw) -> str:
        if raw is True:  return "accepted"
        if raw is False: return "rejected"
        return str(raw)

    csv_path = RESULTS_DIR / analysis_id / "curated_cells.csv"
    if not csv_path.exists():
        try:
            images_data = _load_measurements(analysis_id)
        except HTTPException:
            return None
        rows = []
        for img_data in images_data:
            filename   = img_data["filename"]
            pixel_size = img_data.get("pixel_size_um", 0.1075)
            for cell in img_data.get("cells", []):
                key = _curation_key(analysis_id, filename, cell["cell_id"])
                raw = curation.get(key, True)
                if raw is False:
                    continue
                row = {
                    "strain":        img_data["strain"],
                    "filename":      filename,
                    "pixel_size_um": pixel_size,
                    "filepath":      img_data["filepath"],
                    "morphotype":    _morph(raw),
                }
                row.update(cell)
                rows.append(row)
        return pd.DataFrame(rows) if rows else None

    df = pd.read_csv(csv_path)
    if df.empty:
        return None

    morphs, keep = [], []
    for _, row in df.iterrows():
        key = _curation_key(analysis_id, str(row.get("filename", "")), row.get("cell_id", 0))
        raw = curation.get(key, True)
        morphs.append(_morph(raw))
        keep.append(raw is not False)

    df["morphotype"] = morphs
    import numpy as _np
    filtered = df[_np.array(keep)]
    return filtered if not filtered.empty else None


@app.post("/api/analyses/{analysis_id}/regen-crops")
async def api_regen_crops(analysis_id: str):
    """Rebuild all cell crop PNGs from existing measurements.json without re-running Cellpose.
    Fixes crops that were previously saved as matplotlib figures (old format).
    """
    images_data = _load_measurements(analysis_id)
    if not images_data:
        raise HTTPException(status_code=404, detail="No measurements found for this analysis")

    import tifffile as _tifffile

    def _read_img(path: Path) -> np.ndarray:
        ext = path.suffix.lower()
        if ext in (".jpg", ".jpeg", ".png"):
            from skimage import io as _skio, color as _skcolor
            img = _skio.imread(str(path))
            if img.ndim == 3:
                img = (_skcolor.rgb2gray(img[..., :3]) * 65535)
            return img.astype(np.float32)
        raw = _tifffile.imread(str(path))
        if raw.ndim == 3:
            raw = raw[0] if raw.shape[0] <= 4 else raw[..., 0]
        return raw.astype(np.float32)

    regenerated, skipped, errors = 0, 0, []
    PAD = 40

    for img_data in images_data:
        filepath = img_data.get("filepath", "")
        src = Path(filepath)
        if not src.exists():
            errors.append(f"source not found: {src.name}")
            continue
        try:
            img = _read_img(src)
        except Exception as e:
            errors.append(f"{src.name}: {e}")
            continue

        for cell in img_data.get("cells", []):
            bbox = cell.get("bbox")
            crop_path = cell.get("crop_path", "")
            if not bbox or not crop_path:
                skipped += 1
                continue
            minr, minc, maxr, maxc = bbox
            r0 = max(0, minr - PAD);  c0 = max(0, minc - PAD)
            r1 = min(img.shape[0], maxr + PAD); c1 = min(img.shape[1], maxc + PAD)
            crop = img[r0:r1, c0:c1].astype(np.float32)
            lo, hi = np.percentile(crop, [1, 99])
            crop8 = ((np.clip(crop, lo, hi) - lo) / (hi - lo + 1e-6) * 255).astype(np.uint8)
            out = Path(crop_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            Image.fromarray(crop8, mode="L").save(str(out))
            # Keep feret_crop_path in sync (same raw PNG)
            fcp = cell.get("feret_crop_path", "")
            if fcp and fcp != crop_path:
                Image.fromarray(crop8, mode="L").save(fcp)
            regenerated += 1

    return {"regenerated": regenerated, "skipped": skipped, "errors": errors}


@app.get("/api/results/summary")
def api_results_summary(analysis_id: str = Query(default="default")):
    import pandas as pd
    df = _build_curated_df(analysis_id)
    if df is None or df.empty:
        return []

    _SKIP = {"strain", "filename", "filepath", "cell_id", "pixel_size_um", "morphotype",
             "bbox", "crop_path", "feret_crop_path",
             "centroid_x_px", "centroid_y_px", "orientation_rad",
             "feret_max_angle_rad", "feret_min_angle_rad"}
    measure_cols = [
        c for c in df.columns
        if c not in _SKIP and pd.api.types.is_numeric_dtype(df[c])
    ]

    group_cols = ["strain", "morphotype"] if "morphotype" in df.columns else ["strain"]

    agg_cols: dict = {"n_cells": ("cell_id", "count")}
    for col in measure_cols:
        if col in df.columns:
            agg_cols[f"{col}_mean"] = (col, "mean")
            agg_cols[f"{col}_sd"]   = (col, "std")

    summary = df.groupby(group_cols).agg(**agg_cols).round(3).reset_index()
    return summary.to_dict(orient="records")


@app.get("/api/results/chart-data")
def api_results_chart_data(analysis_id: str = Query(default="default")):
    """Return per-cell measurements grouped by strain + morphotype for charting."""
    images_data = _load_measurements(analysis_id)
    curation = get_curation_state()
    SKIP = {"bbox", "crop_path", "feret_crop_path", "centroid_x_px", "centroid_y_px",
            "orientation_rad", "feret_max_angle_rad", "feret_min_angle_rad", "area_px"}
    METRICS = ["length_um", "breadth_um", "aspect_ratio",
               "feret_max_um", "feret_min_um", "feret_aspect_ratio",
               "area_um2", "solidity", "perimeter_um"]
    rows = []
    for img in images_data:
        for cell in img.get("cells", []):
            key = _curation_key(analysis_id, img["filename"], cell["cell_id"])
            raw = curation.get(key, True)
            morph = "accepted" if raw is True else ("rejected" if raw is False else str(raw))
            row = {"strain": img["strain"], "morphotype": morph}
            for m in METRICS:
                v = cell.get(m)
                if v is not None:
                    row[m] = round(float(v), 4)
            rows.append(row)
    return rows


@app.get("/api/results/download/cells")
def api_results_download_cells(analysis_id: str = Query(default="default")):
    import pandas as pd
    df = _build_curated_df(analysis_id)
    if df is None or df.empty:
        raise HTTPException(status_code=404, detail="No curated data available")
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return Response(
        content=buf.getvalue().encode(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=curated_cells_{analysis_id}.csv"},
    )


@app.get("/api/results/download/excel")
def api_results_download_excel(analysis_id: str = Query(default="default")):
    try:
        out_path = build_excel(analysis_id)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return FileResponse(
        path=str(out_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"measurements_{analysis_id}.xlsx",
    )


@app.post("/api/results/build-excel")
async def api_build_excel(analysis_id: str = Query(default="default")):
    try:
        out_path = build_excel(analysis_id)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"path": str(out_path), "message": "Excel built successfully"}


def build_excel(analysis_id: str) -> Path:
    """Build measurements.xlsx with one tab per strain plus a summary tab."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    import pandas as pd

    csv_path = RESULTS_DIR / analysis_id / "curated_cells.csv"
    if not csv_path.exists():
        raise FileNotFoundError(
            "No curated data. Export CSV from Curate tab first."
        )

    df = pd.read_csv(csv_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="4ADE80")
    summary_header_fill = PatternFill("solid", fgColor="111827")
    summary_header_font = Font(bold=True, color="60A5FA")

    # ── Summary tab ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    numeric_cols = [
        c for c in df.columns
        if c not in ("strain", "filename", "filepath", "cell_id", "pixel_size_um",
                     "bbox", "crop_path", "feret_crop_path",
                     "centroid_x_px", "centroid_y_px", "orientation_rad",
                     "feret_max_angle_rad", "feret_min_angle_rad")
        and pd.api.types.is_numeric_dtype(df[c])
    ]

    sum_cols = ["strain", "n_cells"] + [
        f"{c}_mean" for c in numeric_cols
    ] + [
        f"{c}_sd" for c in numeric_cols
    ]

    for ci, col in enumerate(sum_cols, 1):
        cell = ws_sum.cell(row=1, column=ci, value=col)
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = Alignment(horizontal="center")

    agg_cols: dict = {"n_cells": ("cell_id", "count")}
    for col in numeric_cols:
        if col in df.columns:
            agg_cols[f"{col}_mean"] = (col, "mean")
            agg_cols[f"{col}_sd"]   = (col, "std")

    summary_df = df.groupby("strain").agg(**agg_cols).round(3).reset_index()
    for ri, row in enumerate(summary_df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_sum.cell(row=ri, column=ci, value=val)

    for col in ws_sum.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws_sum.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # ── One tab per strain ────────────────────────────────────────────────────
    drop_cols = ["filepath", "bbox", "crop_path", "feret_crop_path"]

    for strain in sorted(df["strain"].unique()):
        strain_df = df[df["strain"] == strain].drop(
            columns=[c for c in drop_cols if c in df.columns], errors="ignore"
        )
        ws = wb.create_sheet(title=strain[:31])

        cols = list(strain_df.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(strain_df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                # Convert numpy types to Python native for openpyxl
                if hasattr(val, "item"):
                    val = val.item()
                ws.cell(row=ri, column=ci, value=val)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    out_path = RESULTS_DIR / analysis_id / "measurements.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ── Static files & root ───────────────────────────────────────────────────────

@app.get("/")
def root():
    return FileResponse(str(STATIC_DIR / "index.html"))


app.mount("/", StaticFiles(directory=str(STATIC_DIR)), name="static")
